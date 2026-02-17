from __future__ import annotations

"""
Simple Streamlit dashboard for Adivo monitoring.

Features (v1):
- Let consultants run a one-off PubMed search (e.g. "psoriasis AND Risankizumab")
  with an explicit date window (like in PubMed UI).
- Browse existing snapshots and see only competitor-sponsored papers in a table.
- Download the filtered table as a CSV from the browser.

Run with:

  streamlit run src/dashboard/app.py
"""

from dataclasses import asdict
from datetime import date, datetime, timedelta, timezone
from io import StringIO, BytesIO
from typing import List, Optional
import threading

import sys
from pathlib import Path
import difflib

import pandas as pd
import streamlit as st

# Ensure project root is on sys.path so `src.*` imports work when run via Streamlit.
PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.database.db import connect
from src.ingest.pubmed_ingest import one_off_search_with_competitors
from src.ingest.clinicaltrials_ingest import one_off_ctgov_search
from src.reporting.export_snapshot_to_excel import (
    ExportRow,
    fetch_rows_for_snapshot,
    get_latest_snapshot_id,
)
from src.monitoring.seed_molecules import MOLECULES
from src.summarization.generate_summaries import upsert_summaries_for_snapshot


def _get_all_snapshots() -> list[tuple[str, str, Optional[str]]]:
    """Return list of (snapshot_id, created_at, notes) ordered newest first."""
    with connect() as conn:
        rows = conn.execute(
            """
            SELECT snapshot_id, created_at, notes
            FROM snapshots
            ORDER BY created_at DESC
            """
        ).fetchall()
    return [(sid, created_at, notes) for sid, created_at, notes in rows]


def _fetch_all_ctgov_for_snapshot(snapshot_id: str) -> list[dict]:
    """
    Return all ClinicalTrials.gov trials for a snapshot, without competitor filtering.
    Used to let users see the full CT.gov landscape even when no competitors match.
    """
    with connect() as conn:
        rows = conn.execute(
            """
            SELECT
              d.title,
              d.url,
              d.published_date,
              COALESCE(GROUP_CONCAT(DISTINCT a.affiliation_text), '') AS sponsors
            FROM documents d
            LEFT JOIN affiliations a ON a.doc_id = d.doc_id
            WHERE d.snapshot_id = ? AND d.source = 'ctgov'
            GROUP BY d.doc_id, d.title, d.url, d.published_date
            ORDER BY d.published_date DESC, d.doc_id
            """,
            (snapshot_id,),
        ).fetchall()

    result: list[dict] = []
    for title, url, published_date, sponsors in rows:
        result.append(
            {
                "link": url or "",
                "title": title or "",
                "first_posted": published_date or "",
                "sponsors": sponsors or "",
            }
        )
    return result


def _export_rows_to_csv_text(rows: List[ExportRow]) -> str:
    """Serialize rows to CSV text for Streamlit's download button."""
    if not rows:
        return ""
    df = pd.DataFrame([asdict(r) for r in rows])
    buffer = StringIO()
    df.to_csv(buffer, index=False)
    return buffer.getvalue()


def _suggest_molecule_correction(query: str) -> tuple[Optional[str], Optional[str]]:
    """
    Given a PubMed query, try to spot a molecule-like token that is slightly misspelled
    compared to the known MOLECULE names, and return (typed_token, suggested_canonical_name).

    If no good suggestion is found, returns (None, None).
    """
    known_molecules = [name for name, _ in MOLECULES]
    if not known_molecules or not query.strip():
        return None, None

    # Rough tokenisation: split on whitespace and parentheses, ignore boolean operators.
    raw_tokens = (
        query.replace("(", " ")
        .replace(")", " ")
        .replace("[", " ")
        .replace("]", " ")
        .split()
    )
    tokens = [t for t in raw_tokens if t.upper() not in {"AND", "OR", "NOT"}]
    if not tokens:
        return None, None

    # Take the last non-operator token as the likely molecule name.
    candidate = tokens[-1]
    # Strip punctuation like commas or trailing semicolons.
    candidate_clean = candidate.strip(",.;")
    if not candidate_clean:
        return None, None

    match = difflib.get_close_matches(
        candidate_clean, known_molecules, n=1, cutoff=0.8
    )
    if not match:
        return None, None

    suggestion = match[0]
    if suggestion.lower() == candidate_clean.lower():
        return None, None

    return candidate_clean, suggestion


def _run_pubmed_search_ui() -> Optional[str]:
    """
    Sidebar widget to run a one-off PubMed search.
    Returns the new snapshot_id if a search was triggered, else None.
    """
    st.sidebar.header("Run PubMed search")

    # Use a key so the value persists across reruns.
    query = st.sidebar.text_input(
        "PubMed query",
        value=st.session_state.get("pubmed_query", "psoriasis AND Risankizumab"),
        help='Any valid PubMed query, e.g. "psoriasis AND Risankizumab".',
        key="pubmed_query",
    )

    # If the raw query text changed since last run, drop any previous correction.
    last_raw = st.session_state.get("last_raw_query")
    if last_raw != query:
        st.session_state["last_raw_query"] = query
        st.session_state.pop("corrected_pubmed_query", None)

    # Start from either the stored correction or the raw query.
    corrected_query = st.session_state.get("corrected_pubmed_query", query)

    # Light misspelling helper for molecule names, operating on the *raw* query the user typed.
    typed_token, suggestion = _suggest_molecule_correction(query)
    if suggestion:
        st.sidebar.info(f"Did you mean **{suggestion}** instead of **{typed_token}**?")
        if st.sidebar.button(f"Replace '{typed_token}' with '{suggestion}'"):
            corrected_query = query.replace(typed_token, suggestion)
            st.session_state["corrected_pubmed_query"] = corrected_query

    # If we have an active corrected query, show it explicitly.
    if corrected_query != query:
        st.sidebar.success(f"Using corrected query: {corrected_query}")

    today = date.today()
    default_start = today - timedelta(days=30)

    col1, col2 = st.sidebar.columns(2)
    start_date = col1.date_input("Start date", value=default_start)
    end_date = col2.date_input("End date", value=today)

    use_last_24h = st.sidebar.checkbox(
        "Use last 24 hours",
        value=False,
        help="If checked, ignores the date inputs above and searches only for papers from the last 24 hours.",
    )

    retmax = st.sidebar.number_input(
        "Max PMIDs (retmax)",
        min_value=10,
        max_value=500,
        value=200,
        step=10,
    )

    run_btn = st.sidebar.button("Run search", type="primary")

    if not run_btn:
        return None

    # Determine date window
    if use_last_24h:
        end_dt = datetime.now(timezone.utc).date()
        start_dt = end_dt - timedelta(days=1)
    else:
        # Validate dates from inputs
        if start_date > end_date:
            st.sidebar.error("Start date must be on or before end date.")
            return None
        start_dt = start_date
        end_dt = end_date

    mindate = start_dt.strftime("%Y/%m/%d")
    maxdate = end_dt.strftime("%Y/%m/%d")

    with st.spinner("Running PubMed search and tagging competitors…"):
        snapshot_id = one_off_search_with_competitors(
            query=corrected_query,
            mindate=mindate,
            maxdate=maxdate,
            retmax=int(retmax),
        )

    # Start AI summaries in the background so the table appears immediately; summaries appear after refresh.
    key = f"_summary_started_{snapshot_id}"
    if not st.session_state.get(key, False):
        def _run_summaries(sid: str) -> None:
            try:
                upsert_summaries_for_snapshot(snapshot_id=sid, limit=100)
            except Exception:
                pass

        thread = threading.Thread(target=_run_summaries, args=(snapshot_id,), daemon=True)
        thread.start()
        st.session_state[key] = True
        st.sidebar.info("AI summaries are generating in the background. **Refresh the page** in a minute to see them.")
    st.sidebar.success(f"Created snapshot: {snapshot_id}")
    return snapshot_id


def _run_ctgov_search_ui() -> Optional[str]:
    """
    Sidebar: run a ClinicalTrials.gov search by condition (and optional intervention).
    Returns the snapshot_id if a search was triggered, else None.
    """
    st.sidebar.header("Run ClinicalTrials.gov search")

    condition = st.sidebar.text_input(
        "Condition",
        value=st.session_state.get("ctgov_condition", "psoriasis"),
        help="Medical condition to search for, e.g. psoriasis, rheumatoid arthritis.",
        key="ctgov_condition",
    )
    intervention = st.sidebar.text_input(
        "Intervention (optional)",
        value=st.session_state.get("ctgov_intervention", ""),
        help="Drug or intervention name to narrow results.",
        key="ctgov_intervention",
    )
    # Time window based on ClinicalTrials.gov \"first posted\" date.
    today = date.today()
    default_start = today - timedelta(days=30)
    col1, col2 = st.sidebar.columns(2)
    start_date = col1.date_input(
        "Start date (first posted)",
        value=st.session_state.get("ctgov_start_date", default_start),
        key="ctgov_start_date",
    )
    end_date = col2.date_input(
        "End date (first posted)",
        value=st.session_state.get("ctgov_end_date", today),
        key="ctgov_end_date",
    )
    add_to_latest = st.sidebar.checkbox(
        "Add to latest snapshot",
        value=False,
        help="If checked, trials are added to the most recent snapshot; otherwise a new snapshot is created.",
    )
    max_studies = st.sidebar.number_input(
        "Max trials",
        min_value=10,
        max_value=1000,
        value=200,
        step=50,
        key="ctgov_max_studies",
    )

    run_ct_btn = st.sidebar.button("Run CT.gov search", type="primary", key="run_ctgov")

    if not run_ct_btn or not (condition or "").strip():
        return None

    if start_date > end_date:
        st.sidebar.error("Start date must be on or before end date.")
        return None

    snapshot_id: Optional[str] = None
    if add_to_latest:
        snapshot_id = get_latest_snapshot_id()

    with st.spinner("Fetching trials and tagging competitors…"):
        new_sid = one_off_ctgov_search(
            condition=condition.strip(),
            intervention=intervention.strip() or None,
            snapshot_id=snapshot_id,
            max_studies=int(max_studies),
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
        )

    st.sidebar.success(f"ClinicalTrials.gov ingest done. Snapshot: {new_sid}")
    return new_sid


def _snapshot_browser_ui(initial_snapshot_id: Optional[str] = None) -> None:
    """Main area: browse snapshots and show competitor-sponsored papers."""
    st.header("Competitor-sponsored evidence")

    snapshots = _get_all_snapshots()
    if not snapshots:
        st.info("No snapshots found yet. Run a PubMed or ClinicalTrials.gov search in the sidebar first.")
        return

    # Build labels like: "<snapshot_id> (YYYY-MM-DD) - notes"
    options = []
    default_index = 0
    for idx, (sid, created_at, notes) in enumerate(snapshots):
        created_short = created_at.split("T")[0] if created_at else ""
        label_parts = [sid]
        if created_short:
            label_parts.append(f"({created_short})")
        if notes:
            label_parts.append(f"- {notes}")
        label = " ".join(label_parts)
        options.append((label, sid))
        if initial_snapshot_id and sid == initial_snapshot_id:
            default_index = idx

    labels = [lbl for lbl, _ in options]
    selected_label = st.selectbox("Snapshot", labels, index=default_index)
    selected_snapshot_id = dict(options)[selected_label]

    rows: List[ExportRow] = fetch_rows_for_snapshot(selected_snapshot_id)

    # One-click AI summaries for this snapshot (runs in foreground; table updates after).
    st.caption(
        "New searches start AI summaries in the background (refresh in ~1 min to see them). "
        "Results are cached. "
        "Use the button below to generate or re-generate for this snapshot."
    )
    if st.button("Generate AI summaries for this snapshot", type="secondary"):
        try:
            with st.spinner("Generating study and slide summaries…"):
                n = upsert_summaries_for_snapshot(
                    snapshot_id=selected_snapshot_id, limit=100
                )
            st.success(f"Generated summaries for {n} papers.")
            if hasattr(st, "rerun"):
                st.rerun()
            else:
                st.experimental_rerun()
        except Exception as e:
            st.error(f"Summaries could not be generated: {e}. Set ANTHROPIC_API_KEY to enable.")

    st.subheader(f"Competitor-sponsored evidence ({len(rows)})")
    if not rows:
        st.write("No competitor-sponsored evidence found for this snapshot.")
    else:
        df = pd.DataFrame([asdict(r) for r in rows])
        # Make link column clickable
        df_display = df.copy()
        df_display["link"] = df_display["link"].apply(
            lambda url: f"[link]({url})" if url else ""
        )

        st.markdown(
            "Showing only items where sponsor/affiliations match the **COMPETITORS** list "
            "(PubMed papers and/or ClinicalTrials.gov trials).\n\n"
            "- **study_summary**: short summary of design and key results (LLM when available, otherwise abstract).\n"
            "- **slide_summary**: narrative version suitable for pasting into slides.\n"
            "- **source**: PubMed or ClinicalTrials.gov."
        )

        if not df.empty:
            # Narrow columns a bit and let long text wrap inside cells.
            st.dataframe(
                df_display,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "link": st.column_config.TextColumn("link", width=120),
                    "article_title": st.column_config.TextColumn(
                        "article_title",
                        width=260,
                    ),
                    "study_summary": st.column_config.TextColumn(
                        "study_summary",
                        width=380,
                    ),
                    "slide_summary": st.column_config.TextColumn(
                        "slide_summary",
                        width=380,
                    ),
                    "competitors": st.column_config.TextColumn(
                        "competitors",
                        width=160,
                    ),
                    "source": st.column_config.TextColumn(
                        "source",
                        width=100,
                    ),
                },
            )

        if not df.empty:
            # Downloads: CSV and Excel for competitor-sponsored evidence
            csv_text = _export_rows_to_csv_text(rows)
            excel_buffer = BytesIO()
            df.to_excel(excel_buffer, index=False, engine="openpyxl")
            excel_buffer.seek(0)

            # Set column widths and wrap text so Excel shows columns clearly.
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Alignment

                wb = load_workbook(excel_buffer)
                ws = wb.active
                # Column widths: link, article_title, study_summary, slide_summary, competitors, source
                for col, width in enumerate([50, 55, 50, 50, 18, 18], start=1):
                    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
                # Enable text wrap for all cells so long summaries don't overflow
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                excel_buffer_final = BytesIO()
                wb.save(excel_buffer_final)
                excel_buffer_final.seek(0)
            except Exception:
                excel_buffer_final = excel_buffer
                excel_buffer_final.seek(0)

            col_csv, col_xlsx = st.columns(2)
            with col_csv:
                st.download_button(
                    label="Download CSV",
                    data=csv_text,
                    file_name=f"{selected_snapshot_id.replace(':', '-')}.csv",
                    mime="text/csv",
                )
            with col_xlsx:
                st.download_button(
                    label="Download Excel (.xlsx)",
                    data=excel_buffer_final,
                    file_name=f"{selected_snapshot_id.replace(':', '-')}.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                )

    # Always try to show all CT.gov trials, even if no competitors matched.
    ctgov_rows = _fetch_all_ctgov_for_snapshot(selected_snapshot_id)
    st.subheader(f"All ClinicalTrials.gov trials in this snapshot ({len(ctgov_rows)})")
    if not ctgov_rows:
        st.write("No ClinicalTrials.gov trials found for this snapshot.")
    else:
        ct_df = pd.DataFrame(ctgov_rows)
        ct_display = ct_df.copy()
        ct_display["link"] = ct_display["link"].apply(
            lambda url: f"[link]({url})" if url else ""
        )
        st.dataframe(
            ct_display,
            use_container_width=True,
            hide_index=True,
            column_config={
                "link": st.column_config.TextColumn("link", width=120),
                "title": st.column_config.TextColumn("title", width=260),
                "first_posted": st.column_config.TextColumn("first_posted", width=120),
                "sponsors": st.column_config.TextColumn("sponsors", width=260),
            },
        )


def main() -> None:
    st.set_page_config(
        page_title="Adivo Competitor Monitoring",
        layout="wide",
    )
    st.title("Adivo Competitor Monitoring")

    # Left sidebar: run PubMed and/or ClinicalTrials.gov search.
    # Both widgets are always visible; each has its own button.
    new_pubmed_snapshot_id = _run_pubmed_search_ui()
    new_ctgov_snapshot_id = _run_ctgov_search_ui()

    # If both ran in the same interaction, prefer the CT.gov snapshot (more recent),
    # otherwise fall back to whichever ran, else latest existing snapshot.
    snapshot_id = new_ctgov_snapshot_id or new_pubmed_snapshot_id or get_latest_snapshot_id()
    _snapshot_browser_ui(initial_snapshot_id=snapshot_id)


if __name__ == "__main__":
    main()

